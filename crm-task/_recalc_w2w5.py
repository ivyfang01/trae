# -*- coding: utf-8 -*-
"""重算 W2-W5 任务指标 + 异常个案（结合周度明细）。"""
from __future__ import annotations

from collections import defaultdict, Counter
from datetime import datetime, date
from pathlib import Path
import json

import openpyxl

SRC = Path(r"c:\Users\HY\Desktop\大宗酒店费用6.1-7.5日.xlsx")
SUMMARY = Path(r"d:\trae\crm-task\业务员周度明细_2026-06-08_2026-07-05_v2.xlsx")
OUT = Path(r"d:\trae\crm-task\_task_recalc.json")

WEEK = {
    "W2": (date(2026, 6, 8), date(2026, 6, 14)),
    "W3": (date(2026, 6, 15), date(2026, 6, 21)),
    "W4": (date(2026, 6, 22), date(2026, 6, 28)),
    "W5": (date(2026, 6, 29), date(2026, 7, 5)),
}
EXEMPT = {"梁日高", "苏彦驰", "范思聪"}


def to_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if not v:
        return None
    s = str(v)[:19]
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def to_f(v):
    try:
        return float(v or 0)
    except Exception:
        return 0.0


def yes(v):
    return str(v or "").strip() in ("是", "Y", "1", "true", "True")


def week_of(d):
    if not d:
        return None
    for w, (a, b) in WEEK.items():
        if a <= d <= b:
            return w
    return None


def role_bucket(role):
    r = str(role or "").strip()
    if r in ("区域经理", "战区经理", "业务经理"):
        return "区经"
    if r in ("业务代表", "业务主管"):
        return "业务代表"
    return "其他"


def main():
    # ---- 周度明细：角色、费用、申诉、汇总字段 ----
    swb = openpyxl.load_workbook(SUMMARY, data_only=True)
    sws = swb.active
    sh = [c.value for c in next(sws.iter_rows(min_row=1, max_row=1))]
    people = {}
    for row in sws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(sh, row))
        name = d.get("业务员")
        if not name:
            continue
        d["_role"] = role_bucket(d.get("角色"))
        people[str(name).strip()] = d
    swb.close()

    # ---- 任务完成数据 ----
    src = openpyxl.load_workbook(SRC, data_only=True, read_only=True)
    tws = src["6月1号-7.5日任务完成数据"]
    th = None
    tasks = []
    for i, row in enumerate(tws.iter_rows(values_only=True)):
        if i == 0:
            th = {str(c).strip(): j for j, c in enumerate(row) if c}
            continue
        if not row:
            continue
        if not yes(row[th.get("是否创建任务", 4)]):
            continue
        start = to_date(row[th["任务开始时间"]]) if "任务开始时间" in th else None
        w = week_of(start)
        if not w:
            continue
        name = str(row[th.get("员工名字", 0)] or "").strip()
        role = people.get(name, {}).get("_role") or "其他"
        rate = to_f(row[th["任务的完成度"]]) if "任务的完成度" in th else 0
        mod = yes(row[th["是否修改任务"]]) if "是否修改任务" in th else False
        appeal = yes(row[th["是否申诉"]]) if "是否申诉" in th else False
        spend = 0.0
        for k in ("6月1-7.5对应任务的消费金额", "实际消费金额"):
            if k in th:
                spend = to_f(row[th[k]])
                break
        tasks.append(
            {
                "name": name,
                "week": w,
                "role": role,
                "rate": rate,
                "done": rate >= 1,
                "mod": mod,
                "appeal": appeal,
                "spend": spend,
                "task_id": row[th["任务id"]] if "任务id" in th else None,
            }
        )
    src.close()

    creators = {t["name"] for t in tasks}
    n = len(tasks) or 1
    mod_n = sum(1 for t in tasks if t["mod"])
    done_n = sum(1 for t in tasks if t["done"])
    appeal_people = {t["name"] for t in tasks if t["appeal"]}
    # 申诉也看汇总表
    for name, p in people.items():
        if to_f(p.get("申诉次数")) > 0:
            appeal_people.add(name)

    overall = {
        "任务数": len(tasks),
        "建任务人数": len(creators),
        "修改任务数": mod_n,
        "修改率%": round(mod_n / n * 100, 1),
        "100%完成任务数": done_n,
        "100%完成率%": round(done_n / n * 100, 1),
        "未100%完成率%": round((len(tasks) - done_n) / n * 100, 1),
        "申诉人数": len(appeal_people),
        "申诉名单": sorted(appeal_people),
    }

    # 分周
    weeks = {}
    for w in ["W2", "W3", "W4", "W5"]:
        wt = [t for t in tasks if t["week"] == w]
        wn = len(wt) or 1
        weeks[w] = {
            "任务数": len(wt),
            "建任务人数": len({t["name"] for t in wt}),
            "修改率%": round(sum(1 for t in wt if t["mod"]) / wn * 100, 1) if wt else 0,
            "100%完成率%": round(sum(1 for t in wt if t["done"]) / wn * 100, 1) if wt else 0,
        }

    # 分角色（任务窗内有任务的）
    by_role = {}
    for role in ("区经", "业务代表"):
        # 应建：周度明细该角色且非豁免
        elig = [
            n
            for n, p in people.items()
            if p.get("_role") == role and n not in EXEMPT
        ]
        rt = [t for t in tasks if t["name"] in elig]
        creators_r = {t["name"] for t in rt}
        rn = len(rt) or 1
        by_role[role] = {
            "应建人数": len(elig),
            "建任务人数": len(creators_r),
            "建任务率%": round(len(creators_r) / len(elig) * 100, 1) if elig else 0,
            "未建任务": sorted(set(elig) - creators_r),
            "任务数": len(rt),
            "修改率%": round(sum(1 for t in rt if t["mod"]) / rn * 100, 1) if rt else 0,
            "100%完成率%": round(sum(1 for t in rt if t["done"]) / rn * 100, 1) if rt else 0,
        }
        # 分周
        ww = {}
        for w in ["W2", "W3", "W4", "W5"]:
            wt = [t for t in rt if t["week"] == w]
            wn = len(wt) or 1
            ww[w] = {
                "建任务人数": len({t["name"] for t in wt}),
                "建任务率%": round(len({t["name"] for t in wt}) / len(elig) * 100, 1) if elig else 0,
                "修改率%": round(sum(1 for t in wt if t["mod"]) / wn * 100, 1) if wt else 0,
                "100%完成率%": round(sum(1 for t in wt if t["done"]) / wn * 100, 1) if wt else 0,
                "任务数": len(wt),
            }
        by_role[role]["weeks"] = ww

    # ---- 异常个案：结合周度明细 ----
    # 1) 高频修改：考核窗任务≥2 且修改率≥50%
    per = defaultdict(lambda: {"mod": 0, "n": 0, "done": 0, "spend": 0.0})
    for t in tasks:
        per[t["name"]]["n"] += 1
        per[t["name"]]["mod"] += int(t["mod"])
        per[t["name"]]["done"] += int(t["done"])
        per[t["name"]]["spend"] += t["spend"]

    high_mod = []
    for name, v in per.items():
        if v["n"] < 2:
            continue
        rate = v["mod"] / v["n"]
        if rate < 0.5:
            continue
        p = people.get(name, {})
        high_mod.append(
            {
                "业务员": name,
                "角色": p.get("_role") or p.get("角色"),
                "任务数": v["n"],
                "修改次数": v["mod"],
                "修改率": round(rate, 2),
                "100%任务数": v["done"],
                "未完成": v["n"] - v["done"],
                "任务关联消费": round(v["spend"], 2),
                "6月费用": to_f(p.get("6月费用")),
                "vs5月": to_f(p.get("vs5月")),
                "费用变化%": round((to_f(p.get("6月费用")) - to_f(p.get("vs5月"))) / to_f(p.get("vs5月")) * 100, 1)
                if to_f(p.get("vs5月")) > 0
                else None,
                "应扣金额": to_f(p.get("应扣金额")),
                "申诉次数": to_f(p.get("申诉次数")),
            }
        )
    high_mod.sort(key=lambda x: (-x["修改率"], -x["修改次数"]))

    # 2) 有未完成 + 有应扣
    deduct_cases = []
    for name, p in people.items():
        if name in EXEMPT:
            continue
        if to_f(p.get("应扣金额")) > 0 or to_f(p.get("未完成任务数")) > 0:
            if name not in per and to_f(p.get("未完成任务数")) == 0 and to_f(p.get("应扣金额")) == 0:
                continue
            deduct_cases.append(
                {
                    "业务员": name,
                    "角色": p.get("_role") or p.get("角色"),
                    "未完成任务数": to_f(p.get("未完成任务数")),
                    "应扣金额": to_f(p.get("应扣金额")),
                    "申诉次数": to_f(p.get("申诉次数")),
                    "6月费用": to_f(p.get("6月费用")),
                    "vs5月": to_f(p.get("vs5月")),
                    "修改次数": to_f(p.get("修改次数")),
                    "总任务数": to_f(p.get("总任务数")),
                }
            )
    deduct_cases = [c for c in deduct_cases if c["应扣金额"] > 0]
    deduct_cases.sort(key=lambda x: -x["应扣金额"])

    # 3) 费用升>10% 且（高频修改 或 有未完成）
    fee_up_gov = []
    for name, p in people.items():
        if name in EXEMPT:
            continue
        may, june = to_f(p.get("vs5月")), to_f(p.get("6月费用"))
        if may <= 0 or june <= may * 1.1:
            continue
        v = per.get(name)
        mod_rate = (v["mod"] / v["n"]) if v and v["n"] else to_f(p.get("修改次数")) / max(to_f(p.get("总任务数")), 1)
        incomplete = (v["n"] - v["done"]) if v else to_f(p.get("未完成任务数"))
        if mod_rate >= 0.5 or incomplete > 0 or to_f(p.get("应扣金额")) > 0:
            fee_up_gov.append(
                {
                    "业务员": name,
                    "角色": p.get("_role") or p.get("角色"),
                    "vs5月": may,
                    "6月费用": june,
                    "增幅%": round((june - may) / may * 100, 1),
                    "修改率": round(mod_rate, 2) if v else None,
                    "未完成": incomplete,
                    "应扣金额": to_f(p.get("应扣金额")),
                    "治理标签": "费用上升+"
                    + ("高频修改" if mod_rate >= 0.5 else "")
                    + ("/未完成" if incomplete else "")
                    + ("/应扣" if to_f(p.get("应扣金额")) > 0 else ""),
                }
            )
    fee_up_gov.sort(key=lambda x: -x["增幅%"])

    # 4) 未建任务（非豁免）
    no_task = []
    for name, p in people.items():
        if name in EXEMPT:
            continue
        if p.get("_role") not in ("区经", "业务代表"):
            continue
        if name not in creators and str(p.get("是否建任务") or "") != "是":
            no_task.append(
                {
                    "业务员": name,
                    "角色": p.get("_role") or p.get("角色"),
                    "6月费用": to_f(p.get("6月费用")),
                    "vs5月": to_f(p.get("vs5月")),
                    "问题备注": p.get("问题备注"),
                }
            )

    out = {
        "overall": overall,
        "weeks": weeks,
        "by_role": by_role,
        "anomalies": {
            "高频修改": high_mod[:15],
            "应扣金额Top": deduct_cases[:15],
            "费用上升且可治理": fee_up_gov[:15],
            "未建任务非豁免": no_task,
            "应扣合计": round(sum(c["应扣金额"] for c in deduct_cases), 2),
            "高频修改人数": len(high_mod),
            "费用上升可治理人数": len(fee_up_gov),
        },
    }
    OUT.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"overall": overall, "weeks": weeks, "by_role_summary": {k: {a: b for a, b in v.items() if a != "weeks"} for k, v in by_role.items()}, "anomaly_counts": {k: (len(v) if isinstance(v, list) else v) for k, v in out["anomalies"].items() if k != "应扣金额Top" and k != "高频修改" and k != "费用上升且可治理" and k != "未建任务非豁免"} | {"高频修改n": len(high_mod), "应扣n": len(deduct_cases), "费升治理n": len(fee_up_gov), "未建n": len(no_task)}}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
