# -*- coding: utf-8 -*-
"""按《大宗差旅汇报框架》业务员周度明细字段，从桌面费用表+通讯录汇总一人一行 Excel。"""
from __future__ import annotations

from collections import defaultdict
from datetime import datetime, date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

SRC = Path(r"c:\Users\HY\Desktop\大宗酒店费用6.1-7.5日.xlsx")
CONTACT = Path(r"d:\数据查询\输出\7月大宗销售部通讯录.xlsx")
OUT = Path(r"d:\trae\crm-task\业务员周度明细_2026-06-08_2026-07-05_v2.xlsx")

TASK_START = date(2026, 6, 8)
TASK_END = date(2026, 7, 5)
NEW_HIRE_START = date(2026, 6, 1)
NEW_HIRE_END = date(2026, 6, 30)
NEW_MARKET_NAMES = {"岳未郅"}

# 可不建任务 / 不纳入「未建任务异常」的人员说明
EXEMPT_NO_TASK = {
    "梁日高": "5/6月离职，可不建任务",
    "苏彦驰": "5/6月离职，可不建任务",
    "范思聪": "大宗业务负责人，可不建任务",
}

HEADERS = [
    "业务员",
    "区域",
    "站台",
    "新人/老人",
    "新/老市场",
    "是否建任务",
    "客户数",
    "城市数",
    "修改次数",
    "总任务数",
    "未完成任务数",
    "100%完成任务数",
    "任务关联实际酒店消费",
    "6月费用",
    "vs5月",
    "应扣金额",
    "申诉次数",
    "最终扣费",
    "油补是否清零",
    "问题备注",
]


def to_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s[:19], fmt).date()
        except ValueError:
            continue
    return None


def to_float(v):
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def to_int(v):
    if v is None or v == "":
        return 0
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return 0


def yes(v) -> bool:
    return str(v or "").strip() in ("是", "Y", "y", "1", "true", "True")


def split_cities(city_text) -> list[str]:
    if not city_text:
        return []
    s = str(city_text).replace("、", ",").replace("，", ",").replace(";", ",")
    return [p.strip() for p in s.split(",") if p.strip()]


def city_count(city_text) -> int:
    return len(set(split_cities(city_text)))


SKIP_FEE_NAMES = {
    "总计",
    "姓名",
    "名字",
    "(多项)",
    "出行人姓名",
    "出行人名称",
    "完整部门层级",
    "部门",
    "订单类型名称",
}


def normalize_person_names(raw) -> list[str]:
    """拆分汇总表里偶发的「徐帅,刘海」拼接名。"""
    if raw is None:
        return []
    s = str(raw).strip()
    if not s or s in SKIP_FEE_NAMES:
        return []
    parts = []
    for chunk in s.replace("，", ",").replace("、", ",").split(","):
        name = chunk.strip()
        if name and name not in SKIP_FEE_NAMES:
            parts.append(name)
    return parts


def load_roster(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    people = []
    for r in rows[1:]:
        if not r or not r[1]:
            continue
        name = str(r[1]).strip()
        dept = str(r[3]).strip() if r[3] else ""
        hire = to_date(r[12] if len(r) > 12 else None)
        region = str(r[27]).strip() if len(r) > 27 and r[27] else ""
        station = str(r[25]).strip() if len(r) > 25 and r[25] else ""
        if not region and "大宗销售部-" in dept:
            region = dept.split("大宗销售部-")[-1]
        people.append(
            {
                "name": name,
                "region": region,
                "station": station,
                "hire": hire,
                "dept": dept,
                "from_roster": True,
            }
        )
    wb.close()
    return people


def _sheet_header_index(row) -> dict[str, int]:
    return {str(c).strip(): i for i, c in enumerate(row) if c}


def load_consume_hotel(wb, sheet_name: str) -> tuple[dict[str, float], dict[str, set[str]], dict[str, str]]:
    """从「x月大宗消费明细」取：酒店费用、酒店城市、末级部门。"""
    if sheet_name not in wb.sheetnames:
        return {}, {}, {}
    ws = wb[sheet_name]
    fee: dict[str, float] = defaultdict(float)
    cities: dict[str, set[str]] = defaultdict(set)
    region: dict[str, str] = {}
    idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            idx = _sheet_header_index(row)
            continue
        if not row or not idx:
            continue
        ptype = str(row[idx.get("产品类型", 0)] or "").strip()
        if ptype != "酒店":
            continue
        name = str(row[idx.get("预定人", -1)] or "").strip() if "预定人" in idx else ""
        if not name:
            continue
        amt_i = idx.get("合计")
        amt = to_float(row[amt_i]) if amt_i is not None and amt_i < len(row) else 0.0
        fee[name] += amt
        city_i = idx.get("行程/车程/城市")
        if city_i is not None and city_i < len(row) and row[city_i]:
            for c in split_cities(row[city_i]):
                cities[name].add(c)
        if name not in region:
            dept_i = idx.get("预定人末级部门")
            if dept_i is not None and dept_i < len(row) and row[dept_i]:
                region[name] = str(row[dept_i]).strip()
    return (
        {k: round(v, 2) for k, v in fee.items()},
        {k: set(v) for k, v in cities.items()},
        region,
    )


def load_extra_names_from_wb(wb) -> tuple[set[str], dict[str, str]]:
    """补录来源：5/6月大宗消费明细中酒店订单的预定人。"""
    names: set[str] = set()
    region: dict[str, str] = {}
    for sn in ("5月大宗消费明细", "6月大宗消费明细"):
        fee, _cities, reg = load_consume_hotel(wb, sn)
        names |= set(fee.keys())
        for k, v in reg.items():
            region.setdefault(k, v)
    return names, region


def supplement_roster_from_fee_summaries(people: list[dict], wb) -> tuple[list[dict], list[str]]:
    """通讯录 + 消费明细/Sheet11 有而通讯录没有的员工。"""
    existing = {p["name"] for p in people}
    fee_names, region_hints = load_extra_names_from_wb(wb)
    missing = sorted(fee_names - existing)
    for name in missing:
        people.append(
            {
                "name": name,
                "region": region_hints.get(name, ""),
                "station": "",
                "hire": None,
                "dept": "",
                "from_roster": False,
            }
        )
    return people, missing


def load_tasks_from_wb(wb):
    ws = wb["6月1号-7.5日任务完成数据"]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c).strip() if c else "" for c in rows[0]]
    idx = {h: i for i, h in enumerate(header)}

    def get(row, key, default=None):
        i = idx.get(key)
        if i is None or i >= len(row):
            return default
        return row[i]

    by_name = defaultdict(list)
    for row in rows[1:]:
        name = get(row, "员工名字")
        if not name:
            continue
        name = str(name).strip()
        created = yes(get(row, "是否创建任务"))
        start = to_date(get(row, "任务开始时间"))
        if not created:
            by_name[name]
            continue
        if start is None or start < TASK_START or start > TASK_END:
            continue
        modified = yes(get(row, "是否修改任务"))
        cust_before = to_int(get(row, "修改前客户数量"))
        cust_after = to_int(get(row, "修改后客户数量"))
        cust = cust_after if modified and cust_after else cust_before
        hotel_cities = split_cities(get(row, "申请酒店的城市"))
        rate = to_float(get(row, "任务的完成度"))
        # 兼容新旧列名
        spend = to_float(
            get(row, "6月1-7.5对应任务的消费金额")
            or get(row, "实际消费金额")
        )
        appealed = yes(get(row, "是否申诉"))
        by_name[name].append(
            {
                "task_id": get(row, "任务id"),
                "modified": modified,
                "cust": cust,
                "hotel_cities": hotel_cities,
                "rate": rate,
                "spend": spend,
                "appealed": appealed,
            }
        )
    return by_name


def is_new_hire(hire: date | None) -> bool:
    if hire is None:
        return False
    return NEW_HIRE_START <= hire <= NEW_HIRE_END


def aggregate_person(tasks: list[dict], fallback_cities: set[str] | None = None):
    if not tasks:
        cities_n = len(fallback_cities or [])
        return {
            "built": "否",
            "avg_cust": None,
            "hotel_city_n": cities_n,
            "mod_cnt": 0,
            "total": 0,
            "incomplete": 0,
            "done100": 0,
            "task_spend": 0.0,
            "deduct": 0.0,
            "appeal_cnt": 0,
        }
    total = len(tasks)
    mod_cnt = sum(1 for t in tasks if t["modified"])
    incomplete = sum(1 for t in tasks if t["rate"] < 1)
    done100 = sum(1 for t in tasks if t["rate"] >= 1)
    avg_cust = round(sum(t["cust"] for t in tasks) / total, 2)
    hotel_city_n = len({c for t in tasks for c in t["hotel_cities"]})
    if hotel_city_n == 0 and fallback_cities:
        hotel_city_n = len(fallback_cities)
    task_spend = round(sum(t["spend"] for t in tasks), 2)
    deduct = round(
        sum(t["spend"] for t in tasks if t["rate"] < 1 and not t["appealed"]),
        2,
    )
    appeal_cnt = sum(1 for t in tasks if t["appealed"])
    return {
        "built": "是",
        "avg_cust": avg_cust,
        "hotel_city_n": hotel_city_n,
        "mod_cnt": mod_cnt,
        "total": total,
        "incomplete": incomplete,
        "done100": done100,
        "task_spend": task_spend,
        "deduct": deduct,
        "appeal_cnt": appeal_cnt,
    }


def main():
    roster = load_roster(CONTACT)
    before_n = len(roster)

    src_wb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)
    roster, missing = supplement_roster_from_fee_summaries(roster, src_wb)
    # read_only 下任务表用常规打开更稳：关闭后重开非 read_only 仅取任务（文件大时慢）
    # 这里一次性用 read_only 拉任务+费用
    tasks_by = load_tasks_from_wb(src_wb)
    june, june_cities, _ = load_consume_hotel(src_wb, "6月大宗消费明细")
    may, _may_cities, _ = load_consume_hotel(src_wb, "5月大宗消费明细")
    src_wb.close()

    extra_n = len(roster) - before_n
    rows_out = []
    for p in roster:
        name = p["name"]
        agg = aggregate_person(tasks_by.get(name, []), june_cities.get(name))
        june_fee = round(june.get(name, 0.0), 2)
        may_fee = round(may.get(name, 0.0), 2)
        abnormal = agg["built"] == "否" or agg["incomplete"] > 0 or agg["deduct"] > 0
        remark = None
        if name in EXEMPT_NO_TASK:
            remark = EXEMPT_NO_TASK[name]
            # 豁免人员：仅因「未建任务」不标红；其他未完成/应扣仍标
            if agg["built"] == "否" and agg["incomplete"] == 0 and agg["deduct"] == 0:
                abnormal = False
        elif not p.get("from_roster", True):
            remark = "非7月通讯录，来自5/6月大宗消费明细补录"
            abnormal = True
        rows_out.append(
            {
                "业务员": name,
                "区域": p["region"],
                "站台": p["station"],
                "新人/老人": "新人" if is_new_hire(p["hire"]) else "老人",
                "新/老市场": "新市场" if name in NEW_MARKET_NAMES else "老市场",
                "是否建任务": agg["built"],
                "客户数": agg["avg_cust"],
                "城市数": agg["hotel_city_n"],
                "修改次数": agg["mod_cnt"],
                "总任务数": agg["total"],
                "未完成任务数": agg["incomplete"],
                "100%完成任务数": agg["done100"],
                "任务关联实际酒店消费": agg["task_spend"],
                "6月费用": june_fee,
                "vs5月": may_fee,
                "应扣金额": agg["deduct"],
                "申诉次数": agg["appeal_cnt"],
                "最终扣费": None,
                "油补是否清零": None,
                "问题备注": remark,
                "_abnormal": abnormal,
            }
        )

    # 未达标/异常置顶
    rows_out.sort(key=lambda r: (0 if r["_abnormal"] else 1, r["区域"] or "", r["业务员"]))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "业务员周度明细"
    ws.append(HEADERS)

    header_font = Font(bold=True)
    red_font = Font(color="FF0000")
    red_fill = PatternFill("solid", fgColor="FFF2F0")
    thin = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(1, col)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin

    for r in rows_out:
        values = [r[h] for h in HEADERS]
        ws.append(values)
        row_idx = ws.max_row
        for col in range(1, len(HEADERS) + 1):
            cell = ws.cell(row_idx, col)
            cell.border = thin
            if r["_abnormal"]:
                cell.font = red_font
                cell.fill = red_fill

    widths = [10, 10, 12, 10, 10, 10, 8, 8, 8, 8, 10, 12, 14, 10, 10, 10, 8, 8, 10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"

    ws2 = wb.create_sheet("口径说明")
    notes = [
        ["周期", "任务开始时间 2026-06-08 ~ 2026-07-05（含）"],
        ["通讯录", str(CONTACT) + "；并补录5/6月大宗消费明细酒店预定人中通讯录没有的人"],
        ["补录名单", "、".join(missing) if missing else "无"],
        ["数据源", str(SRC)],
        ["新人", "入职日在 2026-06-01~06-30；补录人员无入职日按老人"],
        ["新市场", "仅岳未郅"],
        ["客户数", "按修改后口径；未修改用修改前；对人取任务均值"],
        ["城市数", "优先任务「申请酒店的城市」去重；无任务时用6月消费明细酒店城市"],
        ["修改次数", "是否修改任务=是 的任务数"],
        ["未完成", "完成度 < 1"],
        ["任务关联实际酒店消费", "周期内任务消费金额合计"],
        ["6月费用", "6月大宗消费明细：产品类型=酒店，按预定人汇总「合计」"],
        ["vs5月", "5月大宗消费明细：产品类型=酒店，按预定人汇总「合计」"],
        ["应扣金额", "未完成且未申诉任务的实际酒店消费合计"],
        ["最终扣费/油补/备注", "按框架留空；补录人员备注标明来源"],
        ["标红置顶", "未建任务 或 有未完成任务 或 应扣金额>0 或 消费明细补录"],
    ]
    for row in notes:
        ws2.append(row)
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 80

    wb.save(OUT)
    abnormal_n = sum(1 for r in rows_out if r["_abnormal"])
    print(
        f"ok people={len(rows_out)} extra_from_fee={extra_n} missing={missing} "
        f"abnormal={abnormal_n} built={sum(1 for r in rows_out if r['是否建任务']=='是')} out={OUT}"
    )


if __name__ == "__main__":
    main()
